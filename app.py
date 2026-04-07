import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import copy, io, zipfile, os, re, uuid, traceback
from pathlib import Path

# ══════════════════════════════════════════════════════════════════
# 매핑
# ══════════════════════════════════════════════════════════════════
COMPETENCY_MAP = {
    "Position":     [9, 10, 17, 18, 22, 31],
    "Personality":  [5, 13, 14, 21, 24, 28],
    "Relationship": [6,  7, 23, 25, 30, 32],
    "Results":      [8, 16, 29, 33],
    "Development":  [4, 12, 20, 27],
    "Principles":   [11, 15, 19, 26],
}
SKILL_MAP = {
    "우호성":     [4, 12, 20, 27],
    "동기유발":   [5, 13, 21, 28],
    "자문":       [6, 14],
    "협력제휴":   [7, 15, 22, 29],
    "협상거래":   [8, 16, 23, 30],
    "합리적설득": [9, 17, 24, 31],
    "합법화":     [10, 18, 25, 32],
    "강요":       [11, 19, 26, 33],
}
SOFT_SKILLS = ["우호성", "동기유발", "자문"]
HARD_SKILLS = ["협력제휴", "협상거래", "합리적설득", "합법화", "강요"]
COMP_ROW  = {"Position":4,"Personality":5,"Relationship":6,
              "Results":7,"Development":8,"Principles":9}
SKILL_ROW = {"우호성":12,"동기유발":13,"자문":14,"협력제휴":15,
              "협상거래":16,"합리적설득":17,"합법화":18,"강요":19}

# ══════════════════════════════════════════════════════════════════
# 계산
# ══════════════════════════════════════════════════════════════════
def avg_rows(scores, rows):
    vals = [float(scores.get(str(r-3), 0)) for r in rows]
    return round(sum(vals)/len(vals), 2) if vals else 0.0

def compute(scores):
    c = {k: avg_rows(scores, v) for k, v in COMPETENCY_MAP.items()}
    s = {k: avg_rows(scores, v) for k, v in SKILL_MAP.items()}
    return {"competency": c, "skill_raw": s,
            "soft_avg": round(sum(s[k] for k in SOFT_SKILLS)/3, 2),
            "hard_avg": round(sum(s[k] for k in HARD_SKILLS)/5, 2)}

# ══════════════════════════════════════════════════════════════════
# 파싱 — 새 형태: 개인별 엑셀 대시보드 (A열=문항번호, C열=응답값)
# ══════════════════════════════════════════════════════════════════
def extract_name_from_filename(filename: str) -> str:
    """파일명에서 팀_이름 추출. 예) _사전과제2___LX인터내셔널__영향력_진단_ESS팀_장주민.xlsx → ESS팀_장주민"""
    stem = os.path.splitext(filename)[0]
    parts = [p for p in stem.split('_') if p]
    team_idx = next((i for i, p in enumerate(parts) if '팀' in p or '부' in p or '실' in p), None)
    if team_idx is not None and team_idx + 1 < len(parts):
        return f"{parts[team_idx]}_{parts[team_idx+1]}"
    # fallback: 마지막 두 segment
    if len(parts) >= 2:
        return f"{parts[-2]}_{parts[-1]}"
    return parts[-1] if parts else stem

def parse_single_file(raw: bytes, filename: str) -> dict:
    """개인 엑셀 대시보드 파일 파싱. A열=문항번호(1~30), C열=응답값."""
    name = extract_name_from_filename(filename)
    wb = load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb.active
    scores = {}
    for row in ws.iter_rows(min_row=4, max_row=33, values_only=True):
        q_num = row[0]   # A열: 문항번호
        val   = row[2]   # C열: 응답값
        if isinstance(q_num, int) and isinstance(val, (int, float)):
            scores[str(q_num)] = float(val)
    return {"name": name, "scores": scores}

def parse_multiple_files(uploaded_files) -> list:
    """여러 업로드 파일 → people 리스트"""
    people = []
    errors = []
    for uf in uploaded_files:
        try:
            raw = uf.read()
            person = parse_single_file(raw, uf.name)
            if len(person["scores"]) < 10:
                errors.append(f"⚠️ {uf.name}: 응답값이 너무 적습니다 ({len(person['scores'])}개)")
            else:
                people.append(person)
        except Exception as e:
            errors.append(f"❌ {uf.name}: {e}")
    return people, errors

# ══════════════════════════════════════════════════════════════════
# 엑셀 생성 (기존 로직 유지)
# ══════════════════════════════════════════════════════════════════
def _copy_ws(wb, src, title):
    ws = wb.create_sheet(title=title)
    for cl, cd in src.column_dimensions.items():
        ws.column_dimensions[cl].width = cd.width
    for rn, rd in src.row_dimensions.items():
        ws.row_dimensions[rn].height = rd.height
    for row in src.iter_rows():
        for cell in row:
            nc = ws.cell(row=cell.row, column=cell.column)
            nc.value = cell.value
            if cell.has_style:
                nc.font=copy.copy(cell.font); nc.border=copy.copy(cell.border)
                nc.fill=copy.copy(cell.fill); nc.number_format=cell.number_format
                nc.protection=copy.copy(cell.protection); nc.alignment=copy.copy(cell.alignment)
    for m in src.merged_cells.ranges:
        ws.merge_cells(str(m))
    return ws

def build_excel(people, excel_tpl: bytes) -> bytes:
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    for p in people:
        src = load_workbook(io.BytesIO(excel_tpl)).worksheets[0]
        ws  = _copy_ws(wb, src, p["name"][:31])
        r   = compute(p["scores"])
        ws.cell(1,1).value = p["name"]
        for q in range(1, 31):
            ws.cell(q+3, 3).value = float(p["scores"].get(str(q), 0))
        for k, rl in COMPETENCY_MAP.items():
            avg = r["competency"][k]
            ws.cell(COMP_ROW[k],7).value = round(avg*len(rl),2); ws.cell(COMP_ROW[k],7).number_format="0.00"
            ws.cell(COMP_ROW[k],8).value = avg;                  ws.cell(COMP_ROW[k],8).number_format="0.00"
        for k, rl in SKILL_MAP.items():
            avg = r["skill_raw"][k]
            ws.cell(SKILL_ROW[k],7).value = round(avg*len(rl),2); ws.cell(SKILL_ROW[k],7).number_format="0.00"
            ws.cell(SKILL_ROW[k],8).value = avg;                   ws.cell(SKILL_ROW[k],8).number_format="0.00"
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

# ══════════════════════════════════════════════════════════════════
# PPT 생성 (기존 로직 유지)
# ══════════════════════════════════════════════════════════════════
def _replace_chart_vals(chart_bytes, new_vals):
    s = chart_bytes.decode('utf-8')
    val_m = re.search(r'(<c:val>.*?<c:numCache>)(.*?)(</c:numCache>.*?</c:val>)', s, re.DOTALL)
    if not val_m: return chart_bytes
    before = re.sub(r'<c:ptCount val="\d+"/>', f'<c:ptCount val="{len(new_vals)}"/>', val_m.group(1))
    fmt = re.search(r'<c:formatCode>[^<]*</c:formatCode>', val_m.group(2))
    fmt_tag = fmt.group(0) if fmt else '<c:formatCode>0.00</c:formatCode>'
    pts = ''.join(f'<c:pt idx="{i}"><c:v>{v}</c:v></c:pt>' for i,v in enumerate(new_vals))
    s = (s[:val_m.start()] + before + f'{fmt_tag}<c:ptCount val="{len(new_vals)}"/>{pts}' + val_m.group(3) + s[val_m.end():])
    s = re.sub(
        r'(<c:valAx>.*?<c:scaling>).*?(</c:scaling>)',
        r'\1<c:orientation val="minMax"/><c:max val="5"/><c:min val="0"/>\2',
        s, flags=re.DOTALL
    )
    return s.encode('utf-8')

def _new_guids(s):
    for g in set(re.findall(r'\{[0-9A-F]{8}-[0-9A-F]{4}-[0-9A-F]{4}-[0-9A-F]{4}-[0-9A-F]{12}\}', s)):
        s = s.replace(g, '{'+str(uuid.uuid4()).upper()+'}')
    return s

def _ws_name(n): return f"Microsoft_Excel_Worksheet{n if n>0 else ''}.xlsx"

_PHASE_PLOT_X = 2412488;  _PHASE_BAR_W = 1244294
_STRAT_PLOT_X = 2351917;  _STRAT_BAR_W = 763098
_CIRCLE_PHASE_Y  = 3213000;  _CIRCLE_PHASE_CY  = 437638
_CIRCLE_STRAT_Y  = 6021000;  _CIRCLE_STRAT_CY  = 465643

def _bar_cx_phase(idx):
    return int(_PHASE_PLOT_X + (idx + 0.5) * _PHASE_BAR_W)

def _bar_cx_strat(idx):
    return int(_STRAT_PLOT_X + (idx + 0.5) * _STRAT_BAR_W)

def _update_chart_phase_colors(chart_bytes, vals):
    s = chart_bytes.decode('utf-8')
    max_val = max(vals); min_val = min(vals)
    s = re.sub(r'<c:dPt>.*?</c:dPt>', '', s, flags=re.DOTALL)
    dpts = ''
    if max_val != min_val:
        for idx in [i for i,v in enumerate(vals) if v == max_val]:
            dpts += (f'<c:dPt><c:idx val="{idx}"/><c:invertIfNegative val="0"/><c:bubble3D val="0"/>'
                     f'<c:spPr><a:solidFill><a:srgbClr val="4480B1"/></a:solidFill><a:ln><a:noFill/></a:ln><a:effectLst/></c:spPr></c:dPt>')
        for idx in [i for i,v in enumerate(vals) if v == min_val]:
            dpts += (f'<c:dPt><c:idx val="{idx}"/><c:invertIfNegative val="0"/><c:bubble3D val="0"/>'
                     f'<c:spPr><a:solidFill><a:srgbClr val="C00000"/></a:solidFill><a:ln><a:noFill/></a:ln><a:effectLst/></c:spPr></c:dPt>')
    s = s.replace('<c:dLbls>', dpts + '<c:dLbls>', 1)
    return s.encode('utf-8')

def _update_chart_strategy_colors(chart_bytes, vals):
    s = chart_bytes.decode('utf-8')
    s = re.sub(r'<c:dPt>.*?</c:dPt>', '', s, flags=re.DOTALL)
    s = re.sub(
        r'(<c:spPr>)<a:solidFill>.*?</a:solidFill>',
        r'\1<a:solidFill><a:srgbClr val="FFD000"/></a:solidFill>',
        s, count=1, flags=re.DOTALL
    )
    dpts = ''
    for idx in [3, 9]:
        if idx < len(vals):
            dpts += (
                f'<c:dPt><c:idx val="{idx}"/><c:invertIfNegative val="0"/><c:bubble3D val="0"/>'
                f'<c:spPr><a:solidFill><a:srgbClr val="2D5576"/></a:solidFill><a:ln><a:noFill/></a:ln><a:effectLst/></c:spPr></c:dPt>'
            )
    s = s.replace('<c:dLbls>', dpts + '<c:dLbls>', 1)
    return s.encode('utf-8')

def _move_circle(slide_str, circle_name, new_x, new_y, new_cx, new_cy):
    idx = slide_str.find(f'name="{circle_name}"')
    if idx == -1: return slide_str
    start = slide_str.rfind('<p:pic>', 0, idx)
    end   = slide_str.find('</p:pic>', idx) + len('</p:pic>')
    pic   = slide_str[start:end]
    pic   = re.sub(r'<a:off x="[^"]*" y="[^"]*"/>', f'<a:off x="{new_x}" y="{new_y}"/>', pic)
    pic   = re.sub(r'<a:ext cx="[^"]*" cy="[^"]*"/>', f'<a:ext cx="{new_cx}" cy="{new_cy}"/>', pic)
    return slide_str[:start] + pic + slide_str[end:]

def _get_strat_circle_targets(strat_vals):
    pull_vals = [(i, strat_vals[i]) for i in range(3)]
    push_vals = [(i, strat_vals[i]) for i in range(4, 9)]
    pull_max = max(v for _, v in pull_vals)
    push_max = max(v for _, v in push_vals)
    pull_targets = [i for i, v in pull_vals if v == pull_max]
    push_targets = [i for i, v in push_vals if v == push_max]
    if len(pull_targets) == 3:
        pull_targets = []
    if len(push_targets) >= 3:
        push_targets = []
    return sorted(pull_targets + push_targets)

def _update_circles(slide_str, comp_vals, strat_vals):
    max_val = max(comp_vals); min_val = min(comp_vals)
    max_idx = comp_vals.index(max_val)
    min_idx = comp_vals.index(min_val)
    cw_p = int(_PHASE_BAR_W * 0.85)
    slide_str = _move_circle(slide_str, 'circle2',
        _bar_cx_phase(max_idx) - cw_p//2, _CIRCLE_PHASE_Y, cw_p, _CIRCLE_PHASE_CY)
    slide_str = _move_circle(slide_str, 'circle1',
        _bar_cx_phase(min_idx) - cw_p//2, _CIRCLE_PHASE_Y, cw_p, _CIRCLE_PHASE_CY)
    targets = _get_strat_circle_targets(strat_vals)
    cw_s = int(_STRAT_BAR_W * 0.85)
    OFF_SCREEN = -2057475
    for ci in range(4):
        if ci < len(targets):
            slide_str = _move_circle(slide_str, f'circle{ci+3}',
                _bar_cx_strat(targets[ci]) - cw_s//2, _CIRCLE_STRAT_Y, cw_s, _CIRCLE_STRAT_CY)
        else:
            slide_str = _move_circle(slide_str, f'circle{ci+3}',
                OFF_SCREEN, 3450346, cw_s, _CIRCLE_STRAT_CY)
    return slide_str

def _fill_slide(sl_str, person, result):
    c=result["competency"]; s=result["skill_raw"]; sa=result["soft_avg"]; ha=result["hard_avg"]
    comp_vals  = list(c.values())
    strat_vals = [s[k] for k in SOFT_SKILLS]+[sa]+[s[k] for k in HARD_SKILLS]+[ha]
    sl_str = sl_str.replace("{{NAME}}", person["name"])
    sl_str = _update_circles(sl_str, comp_vals, strat_vals)
    return sl_str

def build_ppt(people, ppt_tpl: bytes) -> bytes:
    with zipfile.ZipFile(io.BytesIO(ppt_tpl)) as src:
        infos = {info.filename: info for info in src.infolist()}
        files = {info.filename: src.read(info.filename) for info in src.infolist()}

    max_chart = max(int(m) for m in re.findall(r'chart(\d+)\.xml', ' '.join(files)))
    max_color = max(int(m) for m in re.findall(r'colors(\d+)\.xml', ' '.join(files)))
    max_style = max(int(m) for m in re.findall(r'style(\d+)\.xml', ' '.join(files)))
    ws_nums   = [int(m) if m else 0 for m in re.findall(r'Worksheet(\d*)\.xlsx', ' '.join(files))]
    max_ws    = max(ws_nums)

    prs_xml  = files["ppt/presentation.xml"]
    prs_rels = files["ppt/_rels/presentation.xml.rels"]
    ct_xml   = files["[Content_Types].xml"]
    max_sid  = max(int(m) for m in re.findall(r'<p:sldId id="(\d+)"', prs_xml.decode()))
    max_rid  = max(int(m) for m in re.findall(r'Id="rId(\d+)"', prs_rels.decode()))

    orig_s2  = files["ppt/slides/slide2.xml"]
    orig_s2r = files["ppt/slides/_rels/slide2.xml.rels"]
    orig_c3  = files["ppt/charts/chart3.xml"]
    orig_c4  = files["ppt/charts/chart4.xml"]
    orig_c3r = files["ppt/charts/_rels/chart3.xml.rels"]
    orig_c4r = files["ppt/charts/_rels/chart4.xml.rels"]

    for i, person in enumerate(people):
        result = compute(person["scores"])
        comp_vals  = list(result["competency"].values())
        strat_vals = ([result["skill_raw"][k] for k in SOFT_SKILLS] + [result["soft_avg"]] +
                      [result["skill_raw"][k] for k in HARD_SKILLS] + [result["hard_avg"]])

        if i == 0:
            sl = _fill_slide(files["ppt/slides/slide1.xml"].decode('utf-8'), person, result)
            files["ppt/slides/slide1.xml"] = sl.encode('utf-8')
            files["ppt/charts/chart1.xml"] = _update_chart_phase_colors(
                _replace_chart_vals(files["ppt/charts/chart1.xml"], comp_vals), comp_vals)
            files["ppt/charts/chart2.xml"] = _update_chart_strategy_colors(
                _replace_chart_vals(files["ppt/charts/chart2.xml"], strat_vals), strat_vals)
        elif i == 1:
            sl = _fill_slide(files["ppt/slides/slide2.xml"].decode('utf-8'), person, result)
            files["ppt/slides/slide2.xml"] = sl.encode('utf-8')
            files["ppt/charts/chart3.xml"] = _update_chart_phase_colors(
                _replace_chart_vals(files["ppt/charts/chart3.xml"], comp_vals), comp_vals)
            files["ppt/charts/chart4.xml"] = _update_chart_strategy_colors(
                _replace_chart_vals(files["ppt/charts/chart4.xml"], strat_vals), strat_vals)
        else:
            sn = i+1
            ca = max_chart+(i-1)*2+1; cb = ca+1
            cola = max_color+(i-1)*2+1; colb = cola+1
            sta  = max_style+(i-1)*2+1; stb  = sta+1
            wsa_n = max_ws+(i-1)*2+1;   wsb_n = wsa_n+1

            sl = _new_guids(orig_s2.decode('utf-8'))
            sl = _fill_slide(sl, person, result)
            files[f"ppt/slides/slide{sn}.xml"] = sl.encode('utf-8')
            files[f"ppt/slides/_rels/slide{sn}.xml.rels"] = (
                orig_s2r
                .replace(b"chart3.xml", f"chart{ca}.xml".encode())
                .replace(b"chart4.xml", f"chart{cb}.xml".encode())
            )
            files[f"ppt/charts/chart{ca}.xml"] = _update_chart_phase_colors(
                _replace_chart_vals(orig_c3, comp_vals), comp_vals)
            files[f"ppt/charts/chart{cb}.xml"] = _update_chart_strategy_colors(
                _replace_chart_vals(orig_c4, strat_vals), strat_vals)
            files[f"ppt/charts/_rels/chart{ca}.xml.rels"] = (
                orig_c3r
                .replace(b"chart3.xml",  f"chart{ca}.xml".encode())
                .replace(b"colors3.xml", f"colors{cola}.xml".encode())
                .replace(b"style3.xml",  f"style{sta}.xml".encode())
                .replace(b"Microsoft_Excel_Worksheet2.xlsx", _ws_name(wsa_n).encode())
            )
            files[f"ppt/charts/_rels/chart{cb}.xml.rels"] = (
                orig_c4r
                .replace(b"chart4.xml",  f"chart{cb}.xml".encode())
                .replace(b"colors4.xml", f"colors{colb}.xml".encode())
                .replace(b"style4.xml",  f"style{stb}.xml".encode())
                .replace(b"Microsoft_Excel_Worksheet3.xlsx", _ws_name(wsb_n).encode())
            )
            files[f"ppt/charts/colors{cola}.xml"] = files["ppt/charts/colors3.xml"]
            files[f"ppt/charts/colors{colb}.xml"] = files["ppt/charts/colors4.xml"]
            files[f"ppt/charts/style{sta}.xml"]   = files["ppt/charts/style3.xml"]
            files[f"ppt/charts/style{stb}.xml"]   = files["ppt/charts/style4.xml"]
            files[f"ppt/embeddings/{_ws_name(wsa_n)}"] = files["ppt/embeddings/Microsoft_Excel_Worksheet2.xlsx"]
            files[f"ppt/embeddings/{_ws_name(wsb_n)}"] = files["ppt/embeddings/Microsoft_Excel_Worksheet3.xlsx"]

            def add_info(nn, rn):
                ni = zipfile.ZipInfo(nn); ni.compress_type = infos[rn].compress_type; infos[nn] = ni
            for nn, rn in [
                (f"ppt/slides/slide{sn}.xml",            "ppt/slides/slide2.xml"),
                (f"ppt/slides/_rels/slide{sn}.xml.rels", "ppt/slides/_rels/slide2.xml.rels"),
                (f"ppt/charts/chart{ca}.xml",            "ppt/charts/chart3.xml"),
                (f"ppt/charts/chart{cb}.xml",            "ppt/charts/chart4.xml"),
                (f"ppt/charts/_rels/chart{ca}.xml.rels", "ppt/charts/_rels/chart3.xml.rels"),
                (f"ppt/charts/_rels/chart{cb}.xml.rels", "ppt/charts/_rels/chart4.xml.rels"),
                (f"ppt/charts/colors{cola}.xml",         "ppt/charts/colors3.xml"),
                (f"ppt/charts/colors{colb}.xml",         "ppt/charts/colors4.xml"),
                (f"ppt/charts/style{sta}.xml",           "ppt/charts/style3.xml"),
                (f"ppt/charts/style{stb}.xml",           "ppt/charts/style4.xml"),
                (f"ppt/embeddings/{_ws_name(wsa_n)}",    "ppt/embeddings/Microsoft_Excel_Worksheet2.xlsx"),
                (f"ppt/embeddings/{_ws_name(wsb_n)}",    "ppt/embeddings/Microsoft_Excel_Worksheet3.xlsx"),
            ]: add_info(nn, rn)

            max_sid+=1; max_rid+=1; rid=f"rId{max_rid}"
            prs_xml  = prs_xml.replace(b'</p:sldIdLst>', f'<p:sldId id="{max_sid}" r:id="{rid}"/></p:sldIdLst>'.encode())
            prs_rels = prs_rels.replace(b'</Relationships>', f'<Relationship Id="{rid}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/slide" Target="slides/slide{sn}.xml"/></Relationships>'.encode())
            ct_xml   = ct_xml.replace(b'</Types>',
                f'<Override PartName="/ppt/slides/slide{sn}.xml" ContentType="application/vnd.openxmlformats-officedocument.presentationml.slide+xml"/>'
                f'<Override PartName="/ppt/charts/chart{ca}.xml" ContentType="application/vnd.openxmlformats-officedocument.drawingml.chart+xml"/>'
                f'<Override PartName="/ppt/charts/chart{cb}.xml" ContentType="application/vnd.openxmlformats-officedocument.drawingml.chart+xml"/>'
                f'</Types>'.encode())

    files["ppt/presentation.xml"]            = prs_xml
    files["ppt/_rels/presentation.xml.rels"] = prs_rels
    files["[Content_Types].xml"]             = ct_xml
    files["docProps/app.xml"] = re.sub(rb'<Slides>\d+</Slides>', f'<Slides>{len(people)}</Slides>'.encode(), files["docProps/app.xml"])

    out = io.BytesIO()
    with zipfile.ZipFile(out, 'w') as zout:
        for name, data in files.items():
            zout.writestr(infos[name], data)
    return out.getvalue()

# ══════════════════════════════════════════════════════════════════
# 템플릿 탐색
# ══════════════════════════════════════════════════════════════════
def find_template(ext: str):
    base = Path(__file__).parent
    found = sorted(base.glob(f"*{ext}"))
    if found: return found[0].read_bytes(), str(found[0])
    found = sorted(Path(os.getcwd()).glob(f"*{ext}"))
    if found: return found[0].read_bytes(), str(found[0])
    return None, None

# ══════════════════════════════════════════════════════════════════
# UI
# ══════════════════════════════════════════════════════════════════
st.set_page_config(page_title="리더십 영향력 진단 결과 자동화", layout="wide")

with st.sidebar:
    st.markdown("### 📋 문항 매핑 참고")
    st.caption("이미지를 클릭하면 확대됩니다")

    def find_image(filename):
        base = Path(__file__).parent
        p = base / filename
        if p.exists(): return str(p)
        p2 = Path(os.getcwd()) / filename
        if p2.exists(): return str(p2)
        return None

    img_stage    = find_image("mapping_stage.jpg")
    img_strategy = find_image("mapping_strategy.jpg")

    if img_stage:
        st.markdown("**▶ 리더십 영향력 단계**")
        st.image(str(img_stage), width='stretch')
    else:
        st.info("mapping_stage.jpg를 GitHub 루트에 업로드해주세요")

    if img_strategy:
        st.markdown("**▶ 리더십 영향력 전략**")
        st.image(str(img_strategy), width='stretch')
    else:
        st.info("mapping_strategy.jpg를 GitHub 루트에 업로드해주세요")

st.title("CLiCK _ 리더십 영향력 진단 결과 자동화")
st.markdown("---")

st.markdown("""
**📋 업로드 전 확인사항**
- 응답자별 엑셀 파일을 **여러 개 동시에 선택**해서 업로드해주세요
- 파일명 형식: `..._팀명_이름.xlsx` (예: `_영향력_진단_ESS팀_장주민.xlsx`)
- 각 파일의 **A열=문항번호(1~30), C열=응답값(1~5)** 형식이어야 합니다
""")

response_files = st.file_uploader(
    "응답자 엑셀 파일들 (.xlsx) — 여러 파일 동시 선택 가능",
    type=["xlsx", "xls"],
    accept_multiple_files=True
)

if response_files:
    st.info(f"📂 {len(response_files)}개 파일 선택됨")
    with st.expander("선택된 파일 목록 확인"):
        for f in response_files:
            parts = [p for p in os.path.splitext(f.name)[0].split('_') if p]
            team_idx = next((i for i, p in enumerate(parts) if '팀' in p or '부' in p or '실' in p), None)
            if team_idx is not None and team_idx + 1 < len(parts):
                label = f"{parts[team_idx]}_{parts[team_idx+1]}"
            elif len(parts) >= 2:
                label = f"{parts[-2]}_{parts[-1]}"
            else:
                label = f.name
            st.write(f"  ✅ `{f.name}` → 시트명: **{label}**")

st.markdown("---")

if st.button("🚀 보고서 생성", type="primary", use_container_width=True):
    if not response_files:
        st.error("❌ 응답자 엑셀 파일을 업로드해주세요."); st.stop()

    excel_tpl, ep = find_template(".xlsx")
    if not excel_tpl: st.error("❌ 엑셀 템플릿 없음 (GitHub 루트에 .xlsx 파일 필요)"); st.stop()

    ppt_tpl, pp = find_template(".pptx")
    if not ppt_tpl: st.error("❌ PPT 템플릿 없음 (GitHub 루트에 .pptx 파일 필요)"); st.stop()

    people, errors = parse_multiple_files(response_files)

    if errors:
        for err in errors:
            st.warning(err)

    if not people:
        st.error("❌ 처리 가능한 응답자가 없습니다."); st.stop()

    with st.spinner(f"⏳ {len(people)}명 보고서 생성 중..."):
        try:
            excel_out = build_excel(people, excel_tpl)
        except Exception as e:
            st.error(f"❌ 엑셀 실패: {e}"); st.code(traceback.format_exc()); st.stop()
        try:
            ppt_out = build_ppt(people, ppt_tpl)
        except Exception as e:
            st.error(f"❌ PPT 실패: {e}"); st.code(traceback.format_exc()); st.stop()

    st.session_state["excel_out"] = excel_out
    st.session_state["ppt_out"]   = ppt_out
    st.session_state["n"]         = len(people)
    st.session_state["names"]     = [p["name"] for p in people]
    st.session_state["done"]      = True

if st.session_state.get("done"):
    excel_out = st.session_state["excel_out"]
    ppt_out   = st.session_state["ppt_out"]
    n         = st.session_state["n"]
    names     = st.session_state["names"]

    st.success(f"🎉 완료: {n}명 처리 → 엑셀 {n}시트 + PPT {n}슬라이드")

    with st.expander("처리된 응답자 목록"):
        for nm in names:
            st.write(f"  ✅ {nm}")

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("CIAM_리더십영향력_진단지.xlsx", excel_out)
        zf.writestr("CIAM_리더십영향력_진단결과.pptx", ppt_out)

    d1, d2, d3 = st.columns(3)
    with d1:
        st.download_button("⬇️ ZIP (전체)", data=zip_buf.getvalue(),
            file_name="CIAM_리더십영향력_결과.zip", mime="application/zip", use_container_width=True)
    with d2:
        st.download_button("⬇️ 엑셀", data=excel_out,
            file_name="CIAM_리더십영향력_진단지.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)
    with d3:
        st.download_button("⬇️ PPT", data=ppt_out,
            file_name="CIAM_리더십영향력_진단결과.pptx",
            mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
            use_container_width=True)
